from __future__ import annotations

import csv
import json
import re
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = Path(r"C:\Users\Admin\Desktop\OEE-2026.xlsx")
MASTER_TS = ROOT / "src" / "data" / "oeeMasterData.generated.ts"
HEADINGS_TS = ROOT / "src" / "data" / "oeeWorkbookHeadings.generated.ts"
MACHINES_CSV = ROOT / "google-sheet" / "machines.csv"
PRODUCTS_CSV = ROOT / "google-sheet" / "product_master.csv"
MINUTES_PER_SLOT = 5
GENERATED_AT = date.today().isoformat()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_number(value: Any) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def display_number(value: float) -> int | float:
    rounded = round(float(value), 2)
    return int(rounded) if rounded.is_integer() else rounded


def slugify(value: str, fallback: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or fallback


def unique_slug(base: str, used: set[str]) -> str:
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def find_indices(headers: list[str], patterns: list[str]) -> list[int]:
    found: list[int] = []
    lowered = [header.lower() for header in headers]
    for index, header in enumerate(lowered):
        if any(pattern.lower() in header for pattern in patterns):
            found.append(index)
    return found


def find_first(headers: list[str], patterns: list[str], default: int = -1) -> int:
    found = find_indices(headers, patterns)
    return found[0] if found else default


def find_last(headers: list[str], patterns: list[str], default: int = -1) -> int:
    found = find_indices(headers, patterns)
    return found[-1] if found else default


def make_id(*parts: str) -> str:
    return slugify("-".join(part for part in parts if part), "item")


def read_sheet_rows(rows_data: list[tuple[Any, ...]]) -> tuple[list[str], list[str], list[str], list[dict[str, Any]], dict[str, int]]:
    width = max((len(row) for row in rows_data), default=0)
    first = list(rows_data[0]) if len(rows_data) > 0 else []
    second = list(rows_data[1]) if len(rows_data) > 1 else []
    third = list(rows_data[2]) if len(rows_data) > 2 else []
    row1 = [clean_text(first[col] if col < len(first) else None) for col in range(width)]
    row2 = [clean_text(second[col] if col < len(second) else None) for col in range(width)]
    headers = [clean_text(third[col] if col < len(third) else None) for col in range(width)]

    date_idx = find_first(headers, ["date", "日期"])
    shift_idx = find_first(headers, ["shift", "班次", "白/夜", "d/n"])
    product_idx = find_first(headers, ["product name", "产品名称"])
    part_idx = find_first(headers, ["part no"])
    if product_idx < 0 and shift_idx >= 0 and part_idx > shift_idx:
        product_idx = shift_idx + 1
    step_idx = find_first(headers, ["step"])
    normal_indices = find_indices(headers, ["normal production", "正常生产", "ormal production"])
    normal_count_idx = normal_indices[0] if normal_indices else -1
    normal_minutes_idx = normal_indices[-1] if normal_indices else -1
    good_idx = find_first(headers, ["good quantity", "合格数"])
    ng_idx = find_first(headers, ["ng quantity", "不合格"])
    test_idx = find_first(headers, ["test", "งานทดสอบ"])
    cavity_idx = find_first(headers, ["quantityof cavities", "quantity of cavities", "模腔"])

    downtime_patterns = {
        "changeoverMinutes": ["cheng shift", "cheng production", "换产"],
        "inspectionMinutes": ["inspection", "检验"],
        "equipmentRepairMinutes": ["equipment repair", "设备维修"],
        "moldRepairMinutes": ["mold repair", "模具维修"],
        "materialChangeMinutes": ["cheng material", "cheng materil", "换料"],
        "emergencyStopMinutes": ["emergency stop", "不明停机"],
        "meetingMinutes": ["meeting", "shot breack", "5s", "换班", "前会"],
        "plannedStopMinutes": ["stop at plan", "计划停机"],
    }
    indexes = {
        "date": date_idx,
        "shift": shift_idx,
        "product": product_idx,
        "part": part_idx,
        "step": step_idx,
        "normalCount": normal_count_idx,
        "normalMinutes": normal_minutes_idx,
        "good": good_idx,
        "ng": ng_idx,
        "test": test_idx,
        "cavity": cavity_idx,
    }
    for key, patterns in downtime_patterns.items():
        candidates = find_indices(headers, patterns)
        if normal_minutes_idx >= 0:
            candidates = [index for index in candidates if index > normal_minutes_idx]
        indexes[key] = candidates[-1] if candidates else -1

    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows_data[3:], start=4):
        values = list(row) + [None] * max(width - len(row), 0)
        production_date = clean_text(values[date_idx]) if date_idx >= 0 else ""
        product_name = clean_text(values[product_idx]) if product_idx >= 0 else ""
        part_no = clean_text(values[part_idx]) if part_idx >= 0 else ""
        if not production_date or not product_name or not part_no:
            continue
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", production_date):
            continue
        rows.append({"rowNumber": row_number, "values": values})

    return row1, row2, headers, rows, indexes


def get_value(values: list[Any], index: int) -> Any:
    return values[index] if 0 <= index < len(values) else None


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    machines: list[dict[str, Any]] = []
    products_map: OrderedDict[str, dict[str, Any]] = OrderedDict()
    seed_logs: list[dict[str, Any]] = []
    headings_sheets: list[dict[str, Any]] = []
    machine_ids: set[str] = set()
    unique_headings: OrderedDict[str, None] = OrderedDict()
    unique_notes: OrderedDict[str, None] = OrderedDict()
    unique_time_codes: OrderedDict[str, None] = OrderedDict()

    created_at = f"{GENERATED_AT}T00:00:00.000Z"

    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        machine_id = unique_slug(slugify(ws.title, f"machine-{sheet_index}"), machine_ids)
        rows_data = list(ws.iter_rows(values_only=True))
        notes, time_codes, headers, rows, idx = read_sheet_rows(rows_data)
        has_step = idx["step"] >= 0
        latest_row = rows[-1]["values"] if rows else []
        capacity_units = clean_number(get_value(latest_row, idx["normalCount"]))
        capacity_minutes = clean_number(get_value(latest_row, idx["normalMinutes"]))
        if capacity_minutes <= 0 and capacity_units > 0:
            capacity_minutes = capacity_units * MINUTES_PER_SLOT
        if capacity_units <= 0 and capacity_minutes > 0:
            capacity_units = capacity_minutes / MINUTES_PER_SLOT

        machines.append(
            {
                "id": machine_id,
                "name": ws.title.strip(),
                "capacityUnits": display_number(capacity_units),
                "capacityMinutes": display_number(capacity_minutes),
                "hasStep": has_step,
                "rowCount": len(rows),
            }
        )

        for value in notes:
            if value:
                unique_notes.setdefault(value, None)
        for value in time_codes:
            if value:
                unique_time_codes.setdefault(value, None)
        for value in headers:
            if value:
                unique_headings.setdefault(value, None)

        headings_sheets.append(
            {
                "sheetName": ws.title.strip(),
                "rowCount": len(rows),
                "columnCount": ws.max_column,
                "hasStep": has_step,
                "timeCodes": [value for value in time_codes if value],
                "notes": [value for value in notes if value],
                "headings": [value for value in headers if value],
            }
        )

        for log_index, row_info in enumerate(rows, start=1):
            values = row_info["values"]
            product_name = clean_text(get_value(values, idx["product"]))
            part_no = clean_text(get_value(values, idx["part"]))
            step = clean_text(get_value(values, idx["step"])) if has_step else "-"
            step = step or "-"
            good = clean_number(get_value(values, idx["good"]))
            ng = clean_number(get_value(values, idx["ng"]))
            test = clean_number(get_value(values, idx["test"])) if idx["test"] >= 0 else 0

            product_key = f"{machine_id}::{product_name}::{part_no}::{step}"
            if product_key not in products_map:
                products_map[product_key] = {
                    "id": make_id(machine_id, product_name, part_no, step),
                    "machineId": machine_id,
                    "machineName": ws.title.strip(),
                    "productName": product_name,
                    "partNo": part_no,
                    "step": step,
                    "sampleGoodQty": 0,
                    "sampleNgQty": 0,
                    "sampleTestQty": 0,
                }
            products_map[product_key]["sampleGoodQty"] += good
            products_map[product_key]["sampleNgQty"] += ng
            products_map[product_key]["sampleTestQty"] += test

            normal_count = clean_number(get_value(values, idx["normalCount"]))
            normal_minutes = clean_number(get_value(values, idx["normalMinutes"]))
            if normal_minutes <= 0 and normal_count > 0:
                normal_minutes = normal_count * MINUTES_PER_SLOT
            cavity = clean_number(get_value(values, idx["cavity"]))
            speed = clean_number(get_value(values, find_first(headers, ["theoretical impulse", "理论冲次"])))

            log: dict[str, Any] = {
                "id": f"seed-{machine_id}-{log_index}",
                "date": clean_text(get_value(values, idx["date"])),
                "shift": clean_text(get_value(values, idx["shift"])),
                "machineId": machine_id,
                "machineName": ws.title.strip(),
                "productName": product_name,
                "partNo": part_no,
                "step": step,
                "workMinutes": display_number(normal_minutes + sum(clean_number(get_value(values, idx[key])) for key in [
                    "changeoverMinutes",
                    "inspectionMinutes",
                    "equipmentRepairMinutes",
                    "moldRepairMinutes",
                    "materialChangeMinutes",
                    "emergencyStopMinutes",
                    "meetingMinutes",
                    "plannedStopMinutes",
                ])),
                "timeSlots": display_number((normal_minutes + sum(clean_number(get_value(values, idx[key])) for key in [
                    "changeoverMinutes",
                    "inspectionMinutes",
                    "equipmentRepairMinutes",
                    "moldRepairMinutes",
                    "materialChangeMinutes",
                    "emergencyStopMinutes",
                    "meetingMinutes",
                    "plannedStopMinutes",
                ])) / MINUTES_PER_SLOT) if normal_minutes else 0,
                "minutesPerSlot": MINUTES_PER_SLOT,
                "machineSpeed": display_number(speed),
                "cavityQty": display_number(cavity),
                "normalMinutes": display_number(normal_minutes),
                "changeoverMinutes": display_number(clean_number(get_value(values, idx["changeoverMinutes"]))),
                "inspectionMinutes": display_number(clean_number(get_value(values, idx["inspectionMinutes"]))),
                "equipmentRepairMinutes": display_number(clean_number(get_value(values, idx["equipmentRepairMinutes"]))),
                "moldRepairMinutes": display_number(clean_number(get_value(values, idx["moldRepairMinutes"]))),
                "materialChangeMinutes": display_number(clean_number(get_value(values, idx["materialChangeMinutes"]))),
                "emergencyStopMinutes": display_number(clean_number(get_value(values, idx["emergencyStopMinutes"]))),
                "meetingMinutes": display_number(clean_number(get_value(values, idx["meetingMinutes"]))),
                "plannedStopMinutes": display_number(clean_number(get_value(values, idx["plannedStopMinutes"]))),
                "goodQty": display_number(good),
                "ngQty": display_number(ng),
                "testQty": display_number(test),
                "note": "",
                "createdAt": created_at,
                "source": "excel-seed",
            }
            seed_logs.append(log)

    products = []
    used_product_ids: set[str] = set()
    for product in products_map.values():
        product = dict(product)
        product["id"] = unique_slug(product["id"], used_product_ids)
        product["sampleGoodQty"] = display_number(product["sampleGoodQty"])
        product["sampleNgQty"] = display_number(product["sampleNgQty"])
        product["sampleTestQty"] = display_number(product["sampleTestQty"])
        products.append(product)

    MASTER_TS.write_text(
        'import type { Machine, ProductMaster, ProductionLog } from "../types";\n\n'
        f"export const machines = {json.dumps(machines, ensure_ascii=False, indent=2)} satisfies Machine[];\n\n"
        f"export const products = {json.dumps(products, ensure_ascii=False, indent=2)} satisfies ProductMaster[];\n\n"
        f"export const seedLogs = {json.dumps(seed_logs, ensure_ascii=False, indent=2)} satisfies ProductionLog[];\n",
        encoding="utf-8",
    )

    headings = {
        "sourceFile": SOURCE_XLSX.name,
        "generatedAt": GENERATED_AT,
        "sheetCount": len(machines),
        "uniqueHeadingCount": len(unique_headings),
        "timeCodes": list(unique_time_codes.keys()),
        "notes": list(unique_notes.keys()),
        "uniqueHeadings": list(unique_headings.keys()),
        "sheets": headings_sheets,
    }
    HEADINGS_TS.write_text(
        "export type OeeWorkbookSheetHeadings = {\n"
        "  sheetName: string;\n"
        "  rowCount: number;\n"
        "  columnCount: number;\n"
        "  hasStep: boolean;\n"
        "  timeCodes: string[];\n"
        "  notes: string[];\n"
        "  headings: string[];\n"
        "};\n\n"
        f"export const oeeWorkbookHeadings = {json.dumps(headings, ensure_ascii=False, indent=2)} satisfies {{\n"
        "  sourceFile: string;\n"
        "  generatedAt: string;\n"
        "  sheetCount: number;\n"
        "  uniqueHeadingCount: number;\n"
        "  timeCodes: string[];\n"
        "  notes: string[];\n"
        "  uniqueHeadings: string[];\n"
        "  sheets: OeeWorkbookSheetHeadings[];\n"
        "};\n",
        encoding="utf-8",
    )

    with MACHINES_CSV.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=["id", "name", "capacityUnits", "capacityMinutes", "hasStep", "rowCount"])
        writer.writeheader()
        writer.writerows([{**machine, "hasStep": "true" if machine["hasStep"] else "false"} for machine in machines])

    with PRODUCTS_CSV.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(
            output,
            fieldnames=["id", "machineId", "machineName", "productName", "partNo", "step", "sampleGoodQty", "sampleNgQty", "sampleTestQty"],
        )
        writer.writeheader()
        writer.writerows(products)

    print(json.dumps({"machines": len(machines), "products": len(products), "seedLogs": len(seed_logs)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
